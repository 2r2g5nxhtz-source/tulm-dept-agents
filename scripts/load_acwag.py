#!/usr/bin/env python3
"""
Загрузка ACWAG.xlsx в railbot_db (таблица acwag_records).

Использование на Hetzner (внутри docker exec или рядом с контейнером):
  python3 scripts/load_acwag.py ACWAG.xlsx \
    "postgresql://railbot:railbot2026@postgres-railway:5432/railbot_db"

Локально:
  PG_CONNECTION_STRING=postgresql://... python3 scripts/load_acwag.py ACWAG.xlsx
"""

import sys
import os
import re
import openpyxl
import psycopg2
from datetime import datetime

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else "ACWAG.xlsx"
PG_CONN   = sys.argv[2] if len(sys.argv) > 2 else os.getenv("PG_CONNECTION_STRING")

CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS acwag_records (
    id          SERIAL PRIMARY KEY,
    year        SMALLINT NOT NULL,
    num         VARCHAR(20),
    wagon_date  DATE,
    company     VARCHAR(200),
    wagon_count INTEGER,
    acwag_code  VARCHAR(50),
    styk        VARCHAR(50),
    baha        NUMERIC(10,2),
    created_at  TIMESTAMP DEFAULT NOW()
);

CREATE INDEX IF NOT EXISTS idx_acwag_company ON acwag_records (LOWER(company));
CREATE INDEX IF NOT EXISTS idx_acwag_year    ON acwag_records (year);
CREATE INDEX IF NOT EXISTS idx_acwag_styk    ON acwag_records (styk);
CREATE INDEX IF NOT EXISTS idx_acwag_date    ON acwag_records (wagon_date);
"""

def parse_count(val):
    """Парсит количество вагонов: '12gr', '12 gr', 12 -> 12"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    m = re.search(r'\d+', str(val))
    return int(m.group()) if m else None

def load_acwag(xlsx_path, pg_conn_str):
    wb   = openpyxl.load_workbook(xlsx_path)
    conn = psycopg2.connect(pg_conn_str)
    cur  = conn.cursor()

    print("Создаём таблицу acwag_records...")
    cur.execute(CREATE_TABLE_SQL)
    conn.commit()

    # Idempotent — можно запускать повторно
    cur.execute("TRUNCATE acwag_records RESTART IDENTITY;")
    conn.commit()

    total = 0
    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        year = int(sheet_name)
        rows_inserted = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            num = row[0]
            if num is None:
                continue

            wagon_date  = row[1].date() if isinstance(row[1], datetime) else None
            company     = str(row[2]).strip() if row[2] else None
            wagon_count = parse_count(row[3])
            acwag_code  = str(row[4]).strip() if row[4] else None
            styk        = str(row[5]).strip() if row[5] else None
            baha        = float(row[6]) if (len(row) > 6 and row[6] is not None) else None

            cur.execute("""
                INSERT INTO acwag_records
                    (year, num, wagon_date, company, wagon_count, acwag_code, styk, baha)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            """, (year, str(num), wagon_date, company, wagon_count,
                  acwag_code, styk, baha))
            rows_inserted += 1

        conn.commit()
        total += rows_inserted
        print(f"  {year}: {rows_inserted} записей загружено")

    cur.close()
    conn.close()
    print(f"\n✅ Итого загружено: {total} записей в acwag_records")

if __name__ == "__main__":
    if not PG_CONN:
        print("❌ Укажи PG_CONNECTION_STRING или передай как 2-й аргумент")
        sys.exit(1)
    load_acwag(XLSX_PATH, PG_CONN)
